from openpyxl import Workbook
from openpyxl.styles import Fill,Border,Alignment,PatternFill,Side



# 定义表头颜色样式为橙色
header_fill = PatternFill('solid', fgColor='FF7F24')
# 定义表中颜色样式为淡黄色
content_fill = PatternFill('solid', fgColor='FFFFE0')
# 定义表尾颜色样式为淡桔红色
bottom_fill = PatternFill('solid', fgColor='EE9572')

# 定义对齐样式横向居中、纵向居中
align = Alignment(horizontal='center', vertical='center')

# 定义边样式为细条
side = Side('thin')
# 定义表头边框样式，有底边和右边
header_border = Border(bottom=side, right=side)
# 定义表中、表尾边框样式，有左边
content_border = Border(left=side)




wb =Workbook()
# 打开工作表
ws = wb.active

# 调整
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15

# 循环第一行单元格，调整表头样式
for cell in ws['A1':'F1'][0]:
    
    print(cell)
    # 设置单元格填充颜色
    cell.fill = header_fill
    # 设置单元格对齐方式
    cell.alignment = align
    # 设置单元格边框
    cell.border = header_border

# 获取最后一行行号
# row_num = ws.max

# 从第二行开始，循环到倒数第二行
for row in ws.iter_rows(min_row=2, max_row=11,min_col=1,max_col=6):
    # 循环取出单元格，调整表中样式
    for cell in row:
        cell.fill = content_fill
        cell.alignment = align
        cell.border = content_border

# 循环最后一行单元格，调整表尾样式
for cell in ws['11']:
    cell.fill = bottom_fill
    cell.alignment = align
    cell.border = content_border

# 保存
wb.save('file.xlsx')




import os
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# 设置目标文件夹路径
path = './部门利润表汇总/'

# 获取文件夹下的所有文件名
file_list = os.listdir(path)
# 遍历文件名列表，取得每一个文件名
for file_name in file_list:
    # 拼接文件路径
    file_path = path + file_name
    print('正在处理：' + file_name)
    # 读取工作簿
    wb = load_workbook(file_path)
    # 读取工作簿中的活跃工作表
    ws = wb.active
  
    # 实例化 LineChart 类，得到 LineChart 对象
    chart = LineChart()
    # 引用工作表的部分数据
    data = Reference(worksheet=ws, min_row=3, max_row=9, min_col=1, max_col=5)
    # 添加被引用的数据到 LineChart 对象
    chart.add_data(data, from_rows=True, titles_from_data=True)
    # 添加 LineChart 对象到工作表中，指定折线图的位置
    ws.add_chart(chart, "C12")

    # 引用工作表的表头数据
    cats = Reference(worksheet=ws, min_row=2, max_row=2, min_col=2, max_col=5)
    # 设置类别轴的标签
    chart.set_categories(cats)
    # 设置 x 轴的标题
    chart.x_axis.title = "季度"
    # 设置 y 轴的标题
    chart.y_axis.title = "利润"
    # 改变线条颜色
    chart.style = 48

    # 保存文件
    wb.save(file_path)
# 在终端提示表格绘图结束
print('恭喜你，工作表中的图绘制成功！')


